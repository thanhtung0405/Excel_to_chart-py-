import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import os
import sys

# Get the directory of the current script
current_script_directory = os.path.dirname(os.path.abspath(__file__))

#input report month
month = input('Introduce month: ')

# Read workbook and select sheet
input_path = os.path.join(current_script_directory, 'MessData_812944.xlsx') #replace name of excel file here

# Check if any files were found
if not input_path:
    print("No files found containing 'MessData' in the current directory.")
    sys.exit(1)
#data_url = r"C:\Users\ttnguyen\Desktop\MessData_321077.csv"
#excel_file = 'Result.xlsx'
data = pd.read_excel(input_path)
output_path = os.path.join(current_script_directory, f'report_{month}.xlsx')

for col in data.columns:
    print(col)
print("Select your desired data: ")
columnn = input()
#value_counts = data[data[columnn].str.contains('ERR')]
value_counts_failed = data[~data[columnn].between(7.5,12)]
grouped_df = value_counts_failed.groupby('MO_no.').size()
grouped_df_total =  data.groupby('MO_no.').size()
grouped_per = (100 -(grouped_df / grouped_df_total *100)).reset_index(name='ERROR per').rename(columns={0: 'ERROR per'})
# Sort DataFrame by the 'MO_no' column in ascending order
sorted_df = grouped_per.sort_values(by='MO_no.', ascending=True)

# Write DataFrame to Excel file
print(sorted_df.shape[0])
sorted_df.to_excel(output_path, index=False)

# Create Excel workbook and load the data
wb = Workbook()
ws = wb.active
# Load data into the worksheet with column headers
for c_idx, col_name in enumerate(sorted_df.columns, start=3):
    ws.cell(row=3, column=c_idx, value=col_name)

for r_idx, row in enumerate(sorted_df.itertuples(index=False), start=4):
    for c_idx, value in enumerate(row, start=3):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Create a bar chart
chart = BarChart()
chart.title = f"Final Yield % of {columnn} in Each (MO_no)"
chart.x_axis.title = "MO_no"
chart.y_axis.title = "Percentage"
chart.style = 5

# Set different colors for each bar
colors = ['FF0000', '00FF00', '0000FF', 'FFFF00', 'FF00FF', '00FFFF']
for i, bar in enumerate(chart.series):
    bar.graphicalProperties.solidFill = colors[i % len(colors)]

# Add data to the chart
data = Reference(ws, min_col=4, min_row=3, max_col=4, max_row= sorted_df.shape[0]+3)
categories = Reference(ws, min_col=3, min_row=4, max_row= sorted_df.shape[0]+3)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)


# Add the chart to the worksheet
ws.add_chart(chart, "H3")

# Add format
title= ws.cell(row= 1, column= 1, value = 'Failed Frequency Report') 
month= ws.cell(row= 2, column= 1, value = month)
title.font = Font('Arial', bold=True, size=20)
month.font = Font('Arial', bold=True, size=10)

# Save the workbook

wb.save(output_path)